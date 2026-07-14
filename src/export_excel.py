import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

TEMPLATE_FILE = r"docs\Template.xlsx"


def fit_image(img, max_width, max_height):

    ratio = min(
        max_width / img.width,
        max_height / img.height
    )

    img.width = int(img.width * ratio)
    img.height = int(img.height * ratio)

    return img


def load_page_plan(image_folder):

    with open(
        Path(image_folder) / "page_plan.json",
        "r"
    ) as f:

        return json.load(f)


def place_half_page(ws, image_folder, section, top=True):

    title_row = 6 if top else 27

    figure_text = (
        f"FIGURE {section['figure_number']:02d}: "
        f"{section['label']}"
    )

    if section["continuation"]:
        figure_text += " (CONT.)"

    ws.cell(
        row=title_row,
        column=4
    ).value = figure_text

    anchors = (
        ["G10", "N10"]
        if top
        else
        ["G31", "N31"]
    )

    for i, filename in enumerate(
        section["images"][:2]
    ):

        image_path = (
            Path(image_folder)
            / filename
        )

        if not image_path.exists():
            continue

        img = Image(str(image_path))

        fit_image(
            img,
            250,
            180
        )

        ws.add_image(
            img,
            anchors[i]
        )


def place_full_page(ws, image_folder, section):

    figure_text = (
        f"FIGURE {section['figure_number']:02d}: "
        f"{section['label']}"
    )

    if section["continuation"]:
        figure_text += " (CONT.)"

    ws.cell(
        row=6,
        column=4
    ).value = figure_text

    anchors = [
        "G10",
        "N10",
        "G18",
        "N18"
    ]

    for i, filename in enumerate(
        section["images"][:4]
    ):

        image_path = (
            Path(image_folder)
            / filename
        )

        if not image_path.exists():
            continue

        img = Image(str(image_path))

        fit_image(
            img,
            240,
            170
        )

        ws.add_image(
            img,
            anchors[i]
        )

    # Remove second figure title
    ws["D27"] = ""


def clear_template_titles(ws):

    ws["D6"] = ""
    ws["D27"] = ""


def main():

    if len(sys.argv) < 2:

        print(
            r"Usage: python src\export_excel.py images"
        )

        sys.exit()

    image_folder = sys.argv[1]

    pages = load_page_plan(
        image_folder
    )

    output_wb = load_workbook(
        TEMPLATE_FILE
    )

    template_sheet = output_wb[
        "Appendix B - Pictures"
    ]

    template_sheet.title = "Page_1"

    first_page = True

    for page in pages:

        if first_page:

            ws = template_sheet
            first_page = False

        else:

            ws = output_wb.copy_worksheet(
                template_sheet
            )

            ws.title = (
                f"Page_{page['page']}"
            )

        clear_template_titles(ws)

        sections = page["sections"]

        if len(sections) == 1:

            section = sections[0]

            if section["full_page"]:

                place_full_page(
                    ws,
                    image_folder,
                    section
                )

            else:

                place_half_page(
                    ws,
                    image_folder,
                    section,
                    top=True
                )

        elif len(sections) == 2:

            place_half_page(
                ws,
                image_folder,
                sections[0],
                top=True
            )

            place_half_page(
                ws,
                image_folder,
                sections[1],
                top=False
            )

    output_file = (
        Path(image_folder)
        / "Prototype_Report.xlsx"
    )

    output_wb.save(
        output_file
    )

    print()
    print("=" * 50)
    print("REPORT GENERATED")
    print("=" * 50)
    print(output_file)
    print("=" * 50)
    print()


if __name__ == "__main__":
    main()