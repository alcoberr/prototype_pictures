import copy
import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image


TEMPLATE = r"docs\Template.xlsx"

PAGE_HEIGHT = 47

TOP_TITLE_ROW = 6
BOTTOM_TITLE_ROW = 27


def load_page_plan(folder):

    with open(
        Path(folder) / "page_plan.json",
        "r"
    ) as f:

        return json.load(f)


def fit_image(img, max_w, max_h):

    ratio = min(
        max_w / img.width,
        max_h / img.height
    )

    img.width = int(
        img.width * ratio
    )

    img.height = int(
        img.height * ratio
    )

    return img


def copy_page_template(ws, page_index):

    if page_index == 0:
        return

    offset = page_index * PAGE_HEIGHT

    for row in range(1, PAGE_HEIGHT + 1):

        target_row = row + offset

        # height
        ws.row_dimensions[
            target_row
        ].height = ws.row_dimensions[
            row
        ].height

        # values / styles
        for col in range(
            1,
            ws.max_column + 1
        ):

            source = ws.cell(
                row=row,
                column=col
            )

            target = ws.cell(
                row=target_row,
                column=col
            )

            if source.value:
                target.value = source.value

            if source.has_style:
                target._style = copy.copy(
                    source._style
                )


def title_row(page_index, top):

    offset = page_index * PAGE_HEIGHT

    if top:
        return TOP_TITLE_ROW + offset

    return BOTTOM_TITLE_ROW + offset


def add_half_page(
    ws,
    page_index,
    section,
    image_folder,
    top=True
):

    row = title_row(
        page_index,
        top
    )

    text = (
        f"FIGURE "
        f"{section['figure_number']:02d}: "
        f"{section['label']}"
    )

    if section["continuation"]:
        text += " (CONT.)"

    ws.cell(
        row=row,
        column=4
    ).value = text

    offset = page_index * PAGE_HEIGHT

    anchors = (
        [
            f"G{10+offset}",
            f"N{10+offset}"
        ]
        if top
        else
        [
            f"G{31+offset}",
            f"N{31+offset}"
        ]
    )

    for i, filename in enumerate(
        section["images"][:2]
    ):

        img = Image(
            str(
                Path(image_folder)
                / filename
            )
        )

        fit_image(
            img,
            250,
            180
        )

        ws.add_image(
            img,
            anchors[i]
        )


def add_full_page(
    ws,
    page_index,
    section,
    image_folder
):

    row = title_row(
        page_index,
        True
    )

    text = (
        f"FIGURE "
        f"{section['figure_number']:02d}: "
        f"{section['label']}"
    )

    if section["continuation"]:
        text += " (CONT.)"

    ws.cell(
        row=row,
        column=4
    ).value = text

    offset = page_index * PAGE_HEIGHT

    anchors = [
        f"G{10+offset}",
        f"N{10+offset}",
        f"G{18+offset}",
        f"N{18+offset}"
    ]

    for i, filename in enumerate(
        section["images"][:4]
    ):

        img = Image(
            str(
                Path(image_folder)
                / filename
            )
        )

        fit_image(
            img,
            230,
            170
        )

        ws.add_image(
            img,
            anchors[i]
        )


def main():

    if len(sys.argv) < 2:

        print(
            r"Usage: python src\export_excel_v2.py images"
        )

        return

    image_folder = sys.argv[1]

    pages = load_page_plan(
        image_folder
    )

    wb = load_workbook(
        TEMPLATE
    )

    ws = wb[
        "Appendix B - Pictures"
    ]

    # build additional pages
    for page_index in range(
        1,
        len(pages)
    ):

        copy_page_template(
            ws,
            page_index
        )

    # render pages
    for page_index, page in enumerate(
        pages
    ):

        sections = page["sections"]

        if len(sections) == 1:

            section = sections[0]

            if section["full_page"]:

                add_full_page(
                    ws,
                    page_index,
                    section,
                    image_folder
                )

            else:

                add_half_page(
                    ws,
                    page_index,
                    section,
                    image_folder,
                    top=True
                )

        else:

            add_half_page(
                ws,
                page_index,
                sections[0],
                image_folder,
                top=True
            )

            add_half_page(
                ws,
                page_index,
                sections[1],
                image_folder,
                top=False
            )

    output = (
        Path(image_folder)
        / "Prototype_Report.xlsx"
    )

    wb.save(output)

    print()
    print("Report generated:")
    print(output)
    print()


if __name__ == "__main__":
    main()
