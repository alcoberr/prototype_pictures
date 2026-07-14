import json
from pathlib import Path
from copy import copy

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.pagebreak import Break

TEMPLATE = r"docs\Template.xlsx"

FIRST_PAGE_HEIGHT = 47
BODY_HEIGHT = 42

BODY_START_ROW = 6
BODY_END_ROW = 47


def load_plan():

    with open(
        Path("images") / "page_plan.json",
        "r"
    ) as f:

        return json.load(f)


def fit_image(img, max_w, max_h):

    ratio = min(
        max_w / img.width,
        max_h / img.height
    )

    img.width = int(
        img.width * ratio
    )

    img.height = int(
        img.height * ratio
    )

    return img


def page_offset(page_index):

    if page_index == 0:
        return 0

    return (
        FIRST_PAGE_HEIGHT +
        ((page_index - 1) * BODY_HEIGHT)
    )


def copy_body(ws, page_index):

    offset = page_offset(
        page_index
    )

    for source_row in range(
        BODY_START_ROW,
        BODY_END_ROW + 1
    ):

        target_row = (
            offset +
            (source_row - BODY_START_ROW + 1)
        )

        ws.row_dimensions[
            target_row
        ].height = (
            ws.row_dimensions[
                source_row
            ].height
        )

        for col in range(
            1,
            ws.max_column + 1
        ):

            src = ws.cell(
                source_row,
                col
            )

            dst = ws.cell(
                target_row,
                col
            )

            if src.value is not None:

                dst.value = src.value

            if src.has_style:

                dst._style = copy(
                    src._style
                )


def add_page_breaks(ws, pages):

    for page_index in range(
        1,
        len(pages)
    ):

        ws.row_breaks.append(
            Break(
                id=page_offset(
                    page_index
                )
            )
        )


def title_row(page_index, top):

    if page_index == 0:

        return (
            6 if top else 27
        )

    base = page_offset(
        page_index
    )

    return (
        base + 1
        if top
        else
        base + 22
    )


def section_title_text(section):

    text = (
        f"FIGURE "
        f"{section['figure_number']:02d}: "
        f"{section['label']}"
    )

    if section["continuation"]:

        text += " (CONT.)"

    return text


def add_image_grid(
    ws,
    image_folder,
    images,
    anchors,
    width,
    height
):

    for i, filename in enumerate(
        images
    ):

        image_path = (
            Path(image_folder)
            / filename
        )

        if not image_path.exists():
            continue

        img = Image(
            str(image_path)
        )

        fit_image(
            img,
            width,
            height
        )

        ws.add_image(
            img,
            anchors[i]
        )


def render_half(
    ws,
    page_index,
    section,
    image_folder,
    top=True
):

    row = title_row(
        page_index,
        top
    )

    ws.cell(
        row=row,
        column=4
    ).value = section_title_text(
        section
    )

    if page_index == 0:

        anchors = (
            ["G10", "N10"]
            if top
            else
            ["G31", "N31"]
        )

    else:

        offset = page_offset(
            page_index
        )

        anchors = (
            [
                f"G{offset+5}",
                f"N{offset+5}"
            ]
            if top
            else
            [
                f"G{offset+26}",
                f"N{offset+26}"
            ]
        )

    add_image_grid(
        ws,
        image_folder,
        section["images"],
        anchors,
        250,
        180
    )


def render_full(
    ws,
    page_index,
    section,
    image_folder
):

    row = title_row(
        page_index,
        True
    )

    ws.cell(
        row=row,
        column=4
    ).value = section_title_text(
        section
    )

    if page_index == 0:

        anchors = [
            "G10",
            "N10",
            "G18",
            "N18"
        ]

    else:

        offset = page_offset(
            page_index
        )

        anchors = [
            f"G{offset+5}",
            f"N{offset+5}",
            f"G{offset+13}",
            f"N{offset+13}"
        ]

    add_image_grid(
        ws,
        image_folder,
        section["images"],
        anchors,
        230,
        170
    )


def main():

    pages = load_plan()

    wb = load_workbook(
        TEMPLATE
    )

    ws = wb[
        "Appendix B - Pictures"
    ]

    # build page bodies
    for page_index in range(
        1,
        len(pages)
    ):

        copy_body(
            ws,
            page_index
        )

    add_page_breaks(
        ws,
        pages
    )

    for page_index, page in enumerate(
        pages
    ):

        sections = page[
            "sections"
        ]

        if len(sections) == 1:

            section = sections[0]

            if section[
                "image_count"
            ] >= 3:

                render_full(
                    ws,
                    page_index,
                    section,
                    "images"
                )

            else:

                render_half(
                    ws,
                    page_index,
                    section,
                    "images",
                    True
                )

        elif len(sections) == 2:

            render_half(
                ws,
                page_index,
                sections[0],
                "images",
                True
            )

            render_half(
                ws,
                page_index,
                sections[1],
                "images",
                False
            )

    output = (
        Path("images")
        / "Prototype_Report.xlsx"
    )

    wb.save(output)

    print()
    print("=" * 60)
    print("REPORT GENERATED")
    print("=" * 60)
    print(output)
    print("=" * 60)
    print()


if __name__ == "__main__":
    main()