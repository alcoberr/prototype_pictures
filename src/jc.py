import json
from pathlib import Path
from copy import copy

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

TEMPLATE_FILE = r"docs\Template.xlsx"
IMAGE_FOLDER = Path("images")


def load_page_plan():

    with open(
        IMAGE_FOLDER / "page_plan.json",
        "r"
    ) as f:

        return json.load(f)


def apply_page_settings(ws):

    ws.print_area = "D1:AQ47"

    try:
        ws.sheet_view.view = "pageBreakPreview"
    except:
        pass

    try:
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
    except:
        pass


def find_placeholders(ws):

    placeholders = {}

    for row in ws.iter_rows():

        for cell in row:

            if not isinstance(cell.value, str):
                continue

            value = cell.value.strip()

            if (
                value.startswith("FIGURE_")
                or
                value.startswith("IMG_")
            ):
                placeholders[value] = cell.coordinate

    return placeholders

def figure_text(section):

    text = (
        f"FIGURE "
        f"{section['figure_number']:02d}: "
        f"{section['label']}"
    )

    if section.get(
        "continuation",
        False
    ):
        text += " (CONT.)"

    return text


def fit_image(img):

    max_width = 900
    max_height = 600

    scale = min(
        max_width / img.width,
        max_height / img.height
    )

    if scale < 1:

        img.width = int(
            img.width * scale
        )

        img.height = int(
            img.height * scale
        )

    return img


def add_image(
    ws,
    image_file,
    anchor
):

    image_path = (
        IMAGE_FOLDER /
        image_file
    )

    print(
        f"{ws.title}: "
        f"{image_file} -> {anchor}"
    )

    if not image_path.exists():

        print(
            "FILE NOT FOUND:",
            image_path
        )

        return

    img = Image(
        str(image_path)
    )

    fit_image(img)

    ws.add_image(
        img,
        anchor
    )

def render_single_section(
    ws,
    placeholders,
    section
):

    print()
    print("RENDERING:", ws.title)
    print("IMAGES:", section["images"])
    print("PLACEHOLDERS:", placeholders)

    if "FIGURE_TOP" in placeholders:

        ws[
            placeholders["FIGURE_TOP"]
        ] = figure_text(section)

    image_slots = []

    for key in (
        "IMG_TOP_1",
        "IMG_TOP_2",
        "IMG_BOTTOM_1",
        "IMG_BOTTOM_2"
    ):

        if key in placeholders:

            image_slots.append(
                placeholders[key]
            )

    print("SLOTS:", image_slots)

    for image_file, slot in zip(
        section["images"],
        image_slots
    ):

        print(
            "ADDING:",
            image_file,
            "->",
            slot
        )

        add_image(
            ws,
            image_file,
            slot
        )


def render_two_sections(
    ws,
    placeholders,
    top,
    bottom
):

    if "FIGURE_TOP" in placeholders:

        ws[
            placeholders["FIGURE_TOP"]
        ] = figure_text(
            top
        )

    if "FIGURE_BOTTOM" in placeholders:

        ws[
            placeholders["FIGURE_BOTTOM"]
        ] = figure_text(
            bottom
        )

    top_slots = []
    bottom_slots = []

    for key in (
        "IMG_TOP_1",
        "IMG_TOP_2"
    ):

        if key in placeholders:

            top_slots.append(
                placeholders[key]
            )

            ws[
                placeholders[key]
            ] = ""

    for key in (
        "IMG_BOTTOM_1",
        "IMG_BOTTOM_2"
    ):

        if key in placeholders:

            bottom_slots.append(
                placeholders[key]
            )

            ws[
                placeholders[key]
            ] = ""

    for image_file, slot in zip(
        top["images"],
        top_slots
    ):

        add_image(
            ws,
            image_file,
            slot
        )

    for image_file, slot in zip(
        bottom["images"],
        bottom_slots
    ):

        add_image(
            ws,
            image_file,
            slot
        )


def render_page(
    ws,
    page
):

    placeholders = find_placeholders(ws)

    sections = page.get(
        "sections",
        []
    )

    print()
    print("=" * 50)
    print("SHEET:", ws.title)
    print("PAGE:", page.get("page"))
    print("SECTIONS:", len(sections))
    print(sections)
    print("=" * 50)

    if len(sections) == 0:

        print(
            f"NO SECTIONS FOUND FOR {ws.title}"
        )

        return

    if len(sections) == 1:

        render_single_section(
            ws,
            placeholders,
            sections[0]
        )

        return

    if len(sections) >= 2:

        render_two_sections(
            ws,
            placeholders,
            sections[0],
            sections[1]
        )

        return

def clone_master(
    wb,
    master,
    title
):

    ws = wb.copy_worksheet(
        master
    )

    ws.title = title

    try:
        ws.print_area = (
            master.print_area
        )
    except:
        pass

    try:
        ws.page_setup = copy(
            master.page_setup
        )
    except:
        pass

    try:
        ws.page_margins = copy(
            master.page_margins
        )
    except:
        pass

    try:
        ws.print_options = copy(
            master.print_options
        )
    except:
        pass

    try:
        ws.sheet_properties = copy(
            master.sheet_properties
        )
    except:
        pass

    try:
        ws.sheet_view = copy(
            master.sheet_view
        )
    except:
        pass

    apply_page_settings(ws)

    return ws


def main():

    pages = load_page_plan()

    wb = load_workbook(
        TEMPLATE_FILE
    )

    first_sheet = wb.worksheets[0]
    master_sheet = wb.worksheets[1]

    report_pages = []

    if pages:

        report_pages.append(
            (
                first_sheet,
                pages[0]
            )
        )

    for page in pages[1:]:

        ws = clone_master(
            wb,
            master_sheet,
            f"Page_{page['page']}"
        )

        print(
            "CREATED:",
            ws.title
        )

        report_pages.append(
            (
                ws,
                page
            )
        )

    print()
    print("WORKBOOK SHEETS")
    print("-" * 50)

    for ws in wb.worksheets:

        print(ws.title)

    print("-" * 50)
    print()

    for ws, page in report_pages:

        print(
            "RENDERING PAGE:",
            page["page"],
            "ON SHEET:",
            ws.title
        )

        render_page(
            ws,
            page
        )

    if master_sheet in wb.worksheets:
        wb.remove(master_sheet)

    output_file = (
        IMAGE_FOLDER /
        "Prototype_Report.xlsx"
    )

    wb.save(
        output_file
    )

    print(
        f"Generated: {output_file}"
    )

if __name__ == "__main__":
    main()