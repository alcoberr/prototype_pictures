from openpyxl import load_workbook

TEMPLATE = r"docs\Template.xlsx"

wb = load_workbook(TEMPLATE)

for ws in wb.worksheets:

    print()
    print("=" * 60)
    print(f"Sheet: {ws.title}")
    print("=" * 60)

    print("\nMerged Cells:")
    for rng in ws.merged_cells.ranges:
        print(f"  {rng}")

    print("\nNon-Empty Cells:")
    for row in ws.iter_rows():

        for cell in row:

            if cell.value is not None:

                print(
                    f"{cell.coordinate}: "
                    f"{repr(cell.value)}"
                )

    print("\nColumn Widths:")

    for col_letter, dim in ws.column_dimensions.items():

        if dim.width:

            print(
                f"{col_letter}: {dim.width}"
            )

    print("\nRow Heights:")

    for row_num, dim in ws.row_dimensions.items():

        if dim.height:

            print(
                f"{row_num}: {dim.height}"
            )

    print("\nImages:")

    if hasattr(ws, "_images"):

        for img in ws._images:

            try:

                print(
                    f"Anchor: "
                    f"{img.anchor._from.col}, "
                    f"{img.anchor._from.row}"
                )

            except Exception:

                print(
                    "Found image."
                )