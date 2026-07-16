from openpyxl import load_workbook

wb = load_workbook(
    r"docs\Template.xlsx"
)

ws = wb["Appendix B - Pictures"]

for row in range(1, 50):

    values = []

    for col in range(1, 45):

        value = ws.cell(row, col).value

        if value is not None:
            values.append(
                f"{row},{col}={value}"
            )

    if values:

        print(
            f"Row {row}:"
        )

        for v in values:
            print("   ", v)

for row in range(49, 100):

    values = []

    for col in range(1, 45):

        value = ws.cell(row, col).value

        if value is not None:
            values.append(str(value))

    if values:
        print(row, values)


from openpyxl import load_workbook

wb = load_workbook(r"docs\Template.xlsx")
ws = wb["Appendix B - Pictures"]

print("Merged Ranges")

for rng in ws.merged_cells.ranges:

    if rng.min_row >= 49:
        print(rng)


from openpyxl import load_workbook

wb = load_workbook(r"docs\Template.xlsx")
ws = wb["Appendix B - Pictures"]

for row in range(1, ws.max_row + 1):

    value = ws.cell(row, 4).value

    if value in (
        "FIGURE_TOP",
        "FIGURE_BOTTOM"
    ):

        print(
            row,
            value
        )