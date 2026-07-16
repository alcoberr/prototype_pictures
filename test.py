# src/find_placeholders.py

from openpyxl import load_workbook

wb = load_workbook(
    r"docs\Template.xlsx"
)

ws = wb["Appendix B - Pictures"]

targets = {
    "FIGURE_TOP",
    "FIGURE_BOTTOM",
    "IMG_TOP_1",
    "IMG_TOP_2",
    "IMG_TOP_3",
    "IMG_TOP_4",
    "IMG_BOTTOM_1",
    "IMG_BOTTOM_2",
    "IMG_BOTTOM_3",
    "IMG_BOTTOM_4",
}

for row in ws.iter_rows():

    for cell in row:

        if cell.value in targets:

            print(
                f"{cell.coordinate}: "
                f"{cell.value}"
            )